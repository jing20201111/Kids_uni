import db
import pandas as pd
import openpyxl as op
import os
import datetime
import getid
from openpyxl import load_workbook
import datetime as dt

def gen_mem_temp(schoolid,filetype):
        # get template object
        basepath  = os.path.dirname(__file__)
        templatePath =os.path.join(basepath,'downloads','End year template.xlsx')
        bg = op.load_workbook(templatePath)
        sheet1 = bg['Sheet1']
        sheet2 = bg['Username and passwords']

        # generating new path 
        cur = db.getCursor()
        cur.execute("SELECT school_name FROM schools WHERE school_id = %s;",(schoolid ,))
        sch_name = cur.fetchone()[0]
        filename = f'{datetime.datetime.now().year}_{sch_name}_{filetype}.xlsx'
        newPath = os.path.join(basepath,'downloads',filename)
        
        # get member information from databse and insert into spreadsheet
        sheet1.cell(column = 4, row =1, value = sch_name.capitalize())
        sheet2.cell(column = 1, row =1, value = sch_name.capitalize())
        sql = "SELECT member_id, first_name, last_name, gender, member_age, ethnicity, \
            continuing_new, status, passport_number, passport_date_issued, \
            ethnicity_info, teaching_research, publication_promos, social_media,username,password FROM \
            members WHERE school_id = %s ORDER BY member_id" % schoolid 
        cur.execute(sql)
        mem_infos = cur.fetchall()
        if mem_infos:
            for i in range(0,len(mem_infos)):
                mem_info = mem_infos[i][0:-2] 
                for j in range(0,len(mem_info)):
                    sheet1.cell(column = j+1,row = 7+i,value = mem_info[j])
                u_p = list(mem_infos[i][1:3])+list(mem_infos[i][-2:])
                for k in range(0,len(u_p)):
                    sheet2.cell(column = k+1, row = 7+i, value = u_p[k])
        # get previous hours info of each member from databse and insert into spreadsheet
        cur.execute("SELECT previous FROM previous WHERE school_id = %s ORDER BY member_id;",(int(schoolid),))
        previous = cur.fetchall()
        for i in range(0,len(previous)):
            sheet1.cell(column=15, row=7+i, value=previous[i][0])


        # get coordinator information from databse and insert into spreadsheet  
        cur.execute("SELECT name, email, phone_number,username, password FROM coordinator \
            WHERE school_id  = %s;",(schoolid ,))
        coor = cur.fetchone()
        for i in range(0,len(coor[0:3])):
            sheet1.cell(column = 4,row = 2 + i,value = coor[0:3][i])
        coor_u_p = [coor[0],'Coordinator']+list(coor[-2:])
        for j in range(0,len(u_p)):
            sheet2.cell(column = j+1, row = 3, value = coor_u_p[j])

        # get events and cut-off date info from databse and insert into spreadsheet

        # for template downloading
        pd_sql_hours = False
        if filetype =='template':
            # insert event info
            cur.execute("SELECT name, event_id FROM events WHERE EXTRACT(YEAR FROM event_date) = \
                EXTRACT(year from now());")
            events = cur.fetchall()
            if events:
                for i in range(0,len(events)):
                    for j in range(0,2):
                        sheet1.cell(column = 23+i, row = 5+j, value = events[i][j])
            pd_sql_event = "SELECT * FROM events WHERE EXTRACT(YEAR FROM event_date) = EXTRACT(year from now()) \
                ORDER BY event_id;"
            # insert data cut-off date
            year = dt.datetime.now().year
            sheet1.cell(column = 4,row = 5,value = f'{year}-12-31')

        # for completed downloading
        if filetype =='completed':
            # insert cut-off date
            cur.execute("SELECT max(year) FROM membershours;")
            date = cur.fetchone()
            sheet1.cell(column = 4,row = 5,value = date[0])
            # insert event title
            cur.execute("SELECT name, event_id FROM events ORDER BY event_id;")
            events = cur.fetchall()
            if events:
                for i in range(0,len(events)):
                    for j in range(0,2):
                        sheet1.cell(column = 23+i, row = 5+j, value = events[i][j])
            
            # get attendance info of each member and insert into spreadsheet
            for i in range(0,len(events)):
                eventid = events[i][1]
                sql = "SELECT attendance.status FROM members LEFT JOIN \
                attendance ON members.member_id = attendance.member_id WHERE members.school_id \
                = %s AND attendance.event_id = %s ORDER BY members.member_id" %(schoolid, eventid)
                cur.execute(sql)
                attends = cur.fetchall()
                for j in range(0,len(attends)):
                    sheet1.cell(column = 23+i, row = 7+j, value = attends[j][0])
            pd_sql_event = "SELECT * FROM events ORDER BY event_id;"
            pd_sql_hours = "SELECT member_id, first_name, last_name, year, term, hours \
            FROM detail_hours WHERE school_id = %s" % int(schoolid)

            # get gown and hat info and insert into spreadsheet
            cur.execute("SELECT gown_size, hat_size FROM members WHERE school_id = %s \
                ORDER BY member_id;",(int(schoolid),))
            results = cur.fetchall()
            for i in range(0,len(results)):
                for j in range(0,2):
                    sheet1.cell(column = 21+i, row = 7+j, value = results[i][j])
            



        #  save new excel file
        bg.save(newPath)

        #  insert new sheet
        book =load_workbook(newPath)
        writer = pd.ExcelWriter(newPath,engine='openpyxl')
        writer.book = book
        # insert new sheet of events info
        df_event = pd.read_sql(pd_sql_event,db.conn)
        df_event.to_excel(writer,index=False,sheet_name='Events')
        # insert new sheet of hours info for completed spreadsheet
        if pd_sql_hours:
            df_hours = pd.read_sql(pd_sql_hours,db.conn)
            df_hours.to_excel(writer,index=False,sheet_name="Hours")
        writer.save()
        # return filename
        return newPath
